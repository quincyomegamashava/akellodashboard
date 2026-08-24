from __future__ import annotations

import io
import json
import os
import shutil
import uuid
from dataclasses import asdict, dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from flask import current_app
from openpyxl import Workbook, load_workbook
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

OUTPUT_COLUMNS = [
    "First name",
    "Last name",
    "Username",
    "Password",
    "Gender",
    "Role",
    "Subject group",
]

MAPPABLE_FIELDS = [
    {"key": "first_name", "label": "First name", "required": True},
    {"key": "last_name", "label": "Last name", "required": True},
    {"key": "username", "label": "Username", "required": True},
    {"key": "password", "label": "Password", "required": True},
    {"key": "gender", "label": "Gender", "required": False},
    {"key": "status", "label": "Status", "required": True},
]

SOURCE_ALIASES = {
    "first_name": ("first name", "firstname", "first_name"),
    "last_name": ("last name", "lastname", "last_name", "surname"),
    "username": ("username", "user name", "user_name"),
    "password": ("password", "pass word", "passcode"),
    "gender": ("gender", "sex"),
    "status": ("status",),
}

DEFAULT_SKIP_SHEETS = {"all students sheet"}
TMP_MAX_AGE = timedelta(hours=24)


class StudentExportError(ValueError):
    """Raised when the uploaded workbook cannot be transformed."""


@dataclass
class ExportSummary:
    sheet_name: str
    exported_rows: int
    exhausted_rows: int


def ensure_schema() -> None:
    from app import db
    from app.models import StudentExportRun

    try:
        db.metadata.create_all(bind=db.engine, tables=[StudentExportRun.__table__])
    except Exception:
        pass


def storage_root() -> Path:
    root = Path(current_app.instance_path) / "student_exports"
    root.mkdir(parents=True, exist_ok=True)
    return root


def tmp_root() -> Path:
    path = storage_root() / "tmp"
    path.mkdir(parents=True, exist_ok=True)
    return path


def _cleanup_tmp() -> None:
    folder = tmp_root()
    cutoff = datetime.utcnow() - TMP_MAX_AGE
    for item in folder.iterdir():
        try:
            mtime = datetime.utcfromtimestamp(item.stat().st_mtime)
            if mtime < cutoff:
                if item.is_dir():
                    shutil.rmtree(item, ignore_errors=True)
                else:
                    item.unlink(missing_ok=True)
        except OSError:
            continue


def save_upload_for_analyze(upload: FileStorage) -> Tuple[str, str]:
    if not upload or not upload.filename:
        raise StudentExportError("Choose an Excel workbook to process.")
    if not upload.filename.lower().endswith(".xlsx"):
        raise StudentExportError("Only .xlsx files are supported.")

    _cleanup_tmp()
    token = uuid.uuid4().hex
    original_filename = os.path.basename(upload.filename)
    dest = tmp_root() / f"{token}.xlsx"
    meta = tmp_root() / f"{token}.json"
    upload.stream.seek(0)
    upload.save(dest)
    meta.write_text(
        json.dumps({"original_filename": original_filename}),
        encoding="utf-8",
    )
    return token, original_filename


def load_staged_upload(token: str) -> Tuple[Path, str]:
    if not token or not all(ch.isalnum() for ch in token):
        raise StudentExportError("The analyzed workbook is no longer available. Upload it again.")
    path = tmp_root() / f"{token}.xlsx"
    meta_path = tmp_root() / f"{token}.json"
    if not path.is_file():
        raise StudentExportError("The analyzed workbook is no longer available. Upload it again.")
    original_filename = f"{token}.xlsx"
    if meta_path.is_file():
        try:
            original_filename = json.loads(meta_path.read_text(encoding="utf-8")).get(
                "original_filename"
            ) or original_filename
        except Exception:
            pass
    return path, original_filename


def discard_staged_upload(token: str) -> None:
    if not token:
        return
    for suffix in (".xlsx", ".json"):
        path = tmp_root() / f"{token}{suffix}"
        try:
            path.unlink(missing_ok=True)
        except OSError:
            pass


def default_display_name(original_filename: str, now: Optional[datetime] = None) -> str:
    now = now or datetime.now()
    stem = Path(original_filename).stem or "student_export"
    return f"{stem}_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"


def resolve_display_name(original_filename: str, override: Optional[str] = None) -> str:
    if override and str(override).strip():
        raw = str(override).strip()
        if not raw.lower().endswith(".xlsx"):
            raw = f"{raw}.xlsx"
        safe = secure_filename(raw) or default_display_name(original_filename)
        if not safe.lower().endswith(".xlsx"):
            safe = f"{safe}.xlsx"
        return safe
    return default_display_name(original_filename)


def persist_run_files(
    *,
    original_path: Path,
    processed_bytes: io.BytesIO,
    original_filename: str,
    display_name: str,
) -> Tuple[str, str, str, str]:
    folder_name = uuid.uuid4().hex
    folder = storage_root() / folder_name
    folder.mkdir(parents=True, exist_ok=True)

    original_stored_name = secure_filename(original_filename) or "original.xlsx"
    if not original_stored_name.lower().endswith(".xlsx"):
        original_stored_name += ".xlsx"
    original_stored_name = f"original_{original_stored_name}"
    processed_stored_name = f"processed_{secure_filename(display_name) or 'export.xlsx'}"
    if not processed_stored_name.lower().endswith(".xlsx"):
        processed_stored_name += ".xlsx"

    original_dest = folder / original_stored_name
    processed_dest = folder / processed_stored_name
    shutil.copyfile(original_path, original_dest)
    processed_bytes.seek(0)
    processed_dest.write_bytes(processed_bytes.read())
    processed_bytes.seek(0)

    rel_original = (Path("student_exports") / folder_name / original_stored_name).as_posix()
    rel_processed = (Path("student_exports") / folder_name / processed_stored_name).as_posix()
    return rel_original, original_stored_name, rel_processed, processed_stored_name


def absolute_stored_path(relative_path: str) -> Path:
    relative = Path(relative_path)
    if relative.is_absolute() or ".." in relative.parts:
        raise StudentExportError("Invalid stored file path.")
    full = (Path(current_app.instance_path) / relative).resolve()
    root = storage_root().resolve()
    if root not in full.parents and full != root:
        raise StudentExportError("Invalid stored file path.")
    return full


def analyze_workbook_path(path: Path) -> dict:
    try:
        workbook = load_workbook(path, data_only=True)
    except Exception as exc:
        raise StudentExportError(f"Could not read workbook: {exc}") from exc

    try:
        workbook = load_workbook(path, data_only=True)
    except Exception as exc:
        raise StudentExportError(f"Could not read workbook: {exc}") from exc

    try:
        sheets = []
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            header_row = _header_row_for_sheet(worksheet)
            headers = _headers_from_row(header_row)
            suggested_mapping = suggest_mapping(headers)
            suggested = _normalize_header(sheet_name) not in DEFAULT_SKIP_SHEETS
            sheets.append(
                {
                    "name": sheet_name,
                    "suggested": suggested,
                    "headers": headers,
                    "suggested_mapping": suggested_mapping,
                }
            )
        return {"sheets": sheets}
    finally:
        workbook.close()


def analyze_upload(upload: FileStorage) -> dict:
    token, original_filename = save_upload_for_analyze(upload)
    path, _ = load_staged_upload(token)
    analysis = analyze_workbook_path(path)
    analysis["token"] = token
    analysis["original_filename"] = original_filename
    analysis["suggested_name"] = default_display_name(original_filename)
    analysis["mappable_fields"] = MAPPABLE_FIELDS
    analysis["output_columns"] = OUTPUT_COLUMNS
    return analysis


def transform_workbook_path(
    path: Path,
    selected_sheets: List[str],
    mapping: Dict[str, Optional[int]],
    month_name: Optional[str] = None,
) -> Tuple[io.BytesIO, List[ExportSummary]]:
    if not selected_sheets:
        raise StudentExportError("Select at least one sheet to process.")

    mapping = _normalize_mapping(mapping)

    try:
        workbook = load_workbook(path, data_only=True)
    except Exception as exc:
        raise StudentExportError(f"Could not read workbook: {exc}") from exc

    try:
        missing_sheets = [name for name in selected_sheets if name not in workbook.sheetnames]
        if missing_sheets:
            raise StudentExportError("Missing selected sheet(s): " + ", ".join(missing_sheets))

        export_rows: List[Dict[str, str]] = []
        summaries: List[ExportSummary] = []
        month_name = month_name or datetime.now().strftime("%B")

        for sheet_name in selected_sheets:
            worksheet = workbook[sheet_name]
            max_col = _sheet_max_column(worksheet)
            rows = worksheet.iter_rows(min_row=1, max_col=max_col, values_only=True)
            header_row = next(rows, None)
            if not header_row:
                raise StudentExportError(f"`{sheet_name}` is empty.")

            headers = _headers_from_row(header_row)
            _validate_mapping_against_headers(sheet_name, headers, mapping)

            exported_rows = 0
            exhausted_rows = 0

            for row_number, row in enumerate(rows, start=2):
                if _row_is_empty(row):
                    continue

                status_value = _parse_status(
                    _cell(row, mapping["status"]),
                    sheet_name=sheet_name,
                    row_number=row_number,
                )
                if status_value == 3:
                    exhausted_rows += 1
                    continue
                remaining_status = status_value - 3
                if remaining_status < 0:
                    raise StudentExportError(
                        f"`{sheet_name}` row {row_number} would produce a negative Status after subtracting 3."
                    )

                learner = {
                    "First name": _required_text(
                        _cell(row, mapping["first_name"]),
                        sheet_name=sheet_name,
                        row_number=row_number,
                        label="First name",
                    ),
                    "Last name": _required_text(
                        _cell(row, mapping["last_name"]),
                        sheet_name=sheet_name,
                        row_number=row_number,
                        label="Last name",
                    ),
                    "Username": _required_text(
                        _cell(row, mapping["username"]),
                        sheet_name=sheet_name,
                        row_number=row_number,
                        label="Username",
                    ),
                    "Password": _required_text(
                        _cell(row, mapping["password"]),
                        sheet_name=sheet_name,
                        row_number=row_number,
                        label="Password",
                    ),
                    "Gender": _optional_text(_cell(row, mapping.get("gender"))),
                    "Role": "student",
                    "Subject group": f"{sheet_name} {month_name}",
                }
                export_rows.append(learner)
                exported_rows += 1

            summaries.append(
                ExportSummary(
                    sheet_name=sheet_name,
                    exported_rows=exported_rows,
                    exhausted_rows=exhausted_rows,
                )
            )

        return _build_output_workbook(export_rows), summaries
    finally:
        workbook.close()


def summaries_as_dicts(summaries: List[ExportSummary]) -> List[dict]:
    return [asdict(item) for item in summaries]


def _sheet_max_column(worksheet) -> int:
    max_col = getattr(worksheet, "max_column", None) or 1
    return max(1, int(max_col))


def _header_row_for_sheet(worksheet):
    max_col = _sheet_max_column(worksheet)
    return next(
        worksheet.iter_rows(min_row=1, max_row=1, max_col=max_col, values_only=True),
        (),
    )


def _headers_from_row(header_row: Iterable[object]) -> List[dict]:
    headers = []
    for index, value in enumerate(header_row):
        label = _optional_text(value)
        if not label:
            label = f"Column {index + 1} (blank)"
        headers.append({"index": index, "label": label, "blank": not bool(_optional_text(value))})
    return headers


def suggest_mapping(headers: List[dict]) -> Dict[str, Optional[int]]:
    header_map: Dict[str, int] = {}
    for item in headers:
        if item.get("blank"):
            continue
        normalized = _normalize_header(item["label"])
        if normalized and normalized not in header_map:
            header_map[normalized] = item["index"]

    suggested: Dict[str, Optional[int]] = {}
    for field in MAPPABLE_FIELDS:
        key = field["key"]
        index = None
        for alias in SOURCE_ALIASES.get(key, ()):
            if alias in header_map:
                index = header_map[alias]
                break
        suggested[key] = index

    if suggested.get("status") is None:
        suggested["status"] = _infer_blank_status_index(headers, header_map)
    return suggested


def _infer_blank_status_index(headers: List[dict], header_map: Dict[str, int]) -> Optional[int]:
    payment_status_index = header_map.get("payment status")
    if payment_status_index is None:
        return None
    adjacent_index = payment_status_index + 1
    for item in headers:
        if item["index"] == adjacent_index and item.get("blank"):
            return adjacent_index
    return None


def _normalize_mapping(mapping: Dict[str, Optional[int]]) -> Dict[str, Optional[int]]:
    if not isinstance(mapping, dict):
        raise StudentExportError("Column mapping is required.")

    resolved: Dict[str, Optional[int]] = {}
    missing = []
    for field in MAPPABLE_FIELDS:
        raw = mapping.get(field["key"], "")
        if raw in ("", None):
            resolved[field["key"]] = None
            if field["required"]:
                missing.append(field["label"])
            continue
        try:
            resolved[field["key"]] = int(raw)
        except (TypeError, ValueError) as exc:
            raise StudentExportError(f"Invalid column mapping for `{field['label']}`.") from exc
    if missing:
        raise StudentExportError("Map these required columns: " + ", ".join(missing) + ".")
    return resolved


def _validate_mapping_against_headers(
    sheet_name: str, headers: List[dict], mapping: Dict[str, Optional[int]]
) -> None:
    max_index = max((item["index"] for item in headers), default=-1)
    for field in MAPPABLE_FIELDS:
        index = mapping.get(field["key"])
        if index is None:
            continue
        if index < 0 or index > max_index:
            raise StudentExportError(
                f"`{sheet_name}` does not have the mapped `{field['label']}` column."
            )


def _cell(row: Iterable[object], index: Optional[int]):
    if index is None:
        return None
    values = list(row) if not isinstance(row, (list, tuple)) else row
    if index < 0 or index >= len(values):
        return None
    return values[index]


def _parse_status(value: object, *, sheet_name: str, row_number: int) -> int:
    if value is None or str(value).strip() == "":
        raise StudentExportError(
            f"`{sheet_name}` row {row_number} has an empty Status value."
        )
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError) as exc:
        raise StudentExportError(
            f"`{sheet_name}` row {row_number} has a non-numeric Status value: {value!r}."
        ) from exc


def _required_text(value: object, *, sheet_name: str, row_number: int, label: str) -> str:
    text = "" if value is None else str(value).strip()
    if not text:
        raise StudentExportError(
            f"`{sheet_name}` row {row_number} is missing `{label}`."
        )
    return text


def _optional_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _row_is_empty(row: Iterable[object]) -> bool:
    return all(value is None or str(value).strip() == "" for value in row)


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().lower().replace("_", " ").split())


def _build_output_workbook(rows: List[Dict[str, str]]) -> io.BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Students"
    sheet.append(OUTPUT_COLUMNS)

    for row in rows:
        sheet.append([row[column] for column in OUTPUT_COLUMNS])

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 32)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
