"""Akello Revenue helpers: privileges, serialization, summary, FY2027 seed."""

from __future__ import annotations

from decimal import Decimal
from functools import wraps
from typing import Any, Dict, List, Optional

from flask import jsonify, request
from flask_login import current_user, login_required

from app import db
from app.blueprints.akello_revenue.models import AkelloRevenueMonth, AkelloRevenuePeriod
from app.models import AppSetting

PRIVILEGE_NAME = "Revenue Reports"

MONTH_NAMES = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}

# Fiscal year display order: March → February
FY_MONTH_ORDER = [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2]

REVENUE_FIELDS = (
    "rev_asl_hlf_usd",
    "rev_asl_hlf_zwl",
    "rev_lib_hlf_usd",
    "rev_lib_hlf_zwl",
    "rev_asl_org_usd",
    "rev_asl_org_zwl",
    "rev_lib_org_usd",
    "rev_lib_org_zwl",
)

SUBSCRIBER_FIELDS = (
    "sub_asl_hlf_usd",
    "sub_asl_hlf_zwl",
    "sub_lib_hlf_usd",
    "sub_lib_hlf_zwl",
    "sub_asl_org_usd",
    "sub_asl_org_zwl",
    "sub_lib_org_usd",
    "sub_lib_org_zwl",
)

# Seeded from Akello Revenue FY27.xlsx (Mar–Jul)
FY2027_SEED_MONTHS: List[Dict[str, Any]] = [
    {
        "month": 3,
        "rev_asl_hlf_usd": 6500,
        "rev_asl_hlf_zwl": 0,
        "rev_lib_hlf_usd": 45000,
        "rev_lib_hlf_zwl": 0,
        "rev_asl_org_usd": 223.35,
        "rev_asl_org_zwl": 185,
        "rev_lib_org_usd": 196.44,
        "rev_lib_org_zwl": 400.01,
        "sub_asl_hlf_usd": 49563,
        "sub_asl_hlf_zwl": 0,
        "sub_lib_hlf_usd": 66186,
        "sub_lib_hlf_zwl": 0,
        "sub_asl_org_usd": 170,
        "sub_asl_org_zwl": 5,
        "sub_lib_org_usd": 36,
        "sub_lib_org_zwl": 6,
    },
    {
        "month": 4,
        "rev_asl_hlf_usd": 3000,
        "rev_asl_hlf_zwl": 0,
        "rev_lib_hlf_usd": 15000,
        "rev_lib_hlf_zwl": 0,
        "rev_asl_org_usd": 221.15,
        "rev_asl_org_zwl": 370,
        "rev_lib_org_usd": 135.74,
        "rev_lib_org_zwl": 59.2,
        "sub_asl_hlf_usd": 6581,
        "sub_asl_hlf_zwl": 0,
        "sub_lib_hlf_usd": 17552,
        "sub_lib_hlf_zwl": 0,
        "sub_asl_org_usd": 165,
        "sub_asl_org_zwl": 8,
        "sub_lib_org_usd": 33,
        "sub_lib_org_zwl": 1,
    },
    {
        "month": 5,
        "rev_asl_hlf_usd": 6354,
        "rev_asl_hlf_zwl": 0,
        "rev_lib_hlf_usd": 15569,
        "rev_lib_hlf_zwl": 0,
        "rev_asl_org_usd": 131,
        "rev_asl_org_zwl": 481,
        "rev_lib_org_usd": 1864,
        "rev_lib_org_zwl": 223,
        "sub_asl_hlf_usd": 6354,
        "sub_asl_hlf_zwl": 0,
        "sub_lib_hlf_usd": 10379,
        "sub_lib_hlf_zwl": 0,
        "sub_asl_org_usd": 85,
        "sub_asl_org_zwl": 2,
        "sub_lib_org_usd": 35,
        "sub_lib_org_zwl": 2,
    },
    {
        "month": 6,
        "rev_asl_hlf_usd": 23334,
        "rev_asl_hlf_zwl": 0,
        "rev_lib_hlf_usd": 774,
        "rev_lib_hlf_zwl": 0,
        "rev_asl_org_usd": 1531,
        "rev_asl_org_zwl": 111,
        "rev_lib_org_usd": 149.8,
        "rev_lib_org_zwl": 297.58,
        "sub_asl_hlf_usd": 2334,
        "sub_asl_hlf_zwl": 0,
        "sub_lib_hlf_usd": 5166,
        "sub_lib_hlf_zwl": 0,
        "sub_asl_org_usd": 78,
        "sub_asl_org_zwl": 3,
        "sub_lib_org_usd": 45,
        "sub_lib_org_zwl": 4,
    },
    {
        "month": 7,
        "rev_asl_hlf_usd": 21113,
        "rev_asl_hlf_zwl": 0,
        "rev_lib_hlf_usd": 35440,
        "rev_lib_hlf_zwl": 0,
        "rev_asl_org_usd": 2309.49,
        "rev_asl_org_zwl": 74,
        "rev_lib_org_usd": 569.78,
        "rev_lib_org_zwl": 135188.56,
        "sub_asl_hlf_usd": 10354,
        "sub_asl_hlf_zwl": 0,
        "sub_lib_hlf_usd": 22415,
        "sub_lib_hlf_zwl": 0,
        "sub_asl_org_usd": 84,
        "sub_asl_org_zwl": 2,
        "sub_lib_org_usd": 72,
        "sub_lib_org_zwl": 730,
    },
]


def _dec(val: Any) -> Decimal:
    if val is None:
        return Decimal("0")
    if isinstance(val, Decimal):
        return val
    return Decimal(str(val))


def _num(val: Any) -> float:
    return float(_dec(val))


def _int(val: Any) -> int:
    try:
        return int(val or 0)
    except (TypeError, ValueError):
        return 0


def can_view_akello_revenue() -> bool:
    if not current_user.is_authenticated:
        return False
    role = (getattr(current_user, "userRole", None) or "").strip()
    if role == "Admin":
        return True
    try:
        return bool(current_user.has_privilege(PRIVILEGE_NAME))
    except Exception:
        return False


def can_edit_akello_revenue() -> bool:
    if not current_user.is_authenticated:
        return False
    return (getattr(current_user, "userRole", None) or "").strip() == "Admin"


def view_required(f):
    @wraps(f)
    @login_required
    def wrapped(*args, **kwargs):
        if not can_view_akello_revenue():
            return jsonify({"error": "Unauthorized"}), 403
        return f(*args, **kwargs)

    return wrapped


def edit_required(f):
    @wraps(f)
    @login_required
    def wrapped(*args, **kwargs):
        if not can_edit_akello_revenue():
            return jsonify({"error": "Unauthorized"}), 403
        return f(*args, **kwargs)

    return wrapped


def get_zig_rate(period: Optional[AkelloRevenuePeriod] = None) -> Decimal:
    if period is not None and period.zig_usd_rate is not None:
        rate = _dec(period.zig_usd_rate)
        if rate > 0:
            return rate
    try:
        raw = AppSetting.get_value("revenue_reports_zig_exchange", "37") or "37"
        rate = _dec(raw)
        if rate > 0:
            return rate
    except Exception:
        pass
    return Decimal("37")


def month_to_dict(row: AkelloRevenueMonth) -> Dict[str, Any]:
    return {
        "id": row.id,
        "month": row.month,
        "month_name": MONTH_NAMES.get(row.month, str(row.month)),
        "rev_asl_hlf_usd": _num(row.rev_asl_hlf_usd),
        "rev_asl_hlf_zwl": _num(row.rev_asl_hlf_zwl),
        "rev_lib_hlf_usd": _num(row.rev_lib_hlf_usd),
        "rev_lib_hlf_zwl": _num(row.rev_lib_hlf_zwl),
        "rev_asl_org_usd": _num(row.rev_asl_org_usd),
        "rev_asl_org_zwl": _num(row.rev_asl_org_zwl),
        "rev_lib_org_usd": _num(row.rev_lib_org_usd),
        "rev_lib_org_zwl": _num(row.rev_lib_org_zwl),
        "sub_asl_hlf_usd": _int(row.sub_asl_hlf_usd),
        "sub_asl_hlf_zwl": _int(row.sub_asl_hlf_zwl),
        "sub_lib_hlf_usd": _int(row.sub_lib_hlf_usd),
        "sub_lib_hlf_zwl": _int(row.sub_lib_hlf_zwl),
        "sub_asl_org_usd": _int(row.sub_asl_org_usd),
        "sub_asl_org_zwl": _int(row.sub_asl_org_zwl),
        "sub_lib_org_usd": _int(row.sub_lib_org_usd),
        "sub_lib_org_zwl": _int(row.sub_lib_org_zwl),
        "updated_at": row.updated_at.isoformat() if row.updated_at else None,
    }


def period_to_dict(period: AkelloRevenuePeriod, include_months: bool = False) -> Dict[str, Any]:
    data: Dict[str, Any] = {
        "id": period.id,
        "code": period.code,
        "name": period.name,
        "zig_usd_rate": _num(period.zig_usd_rate) if period.zig_usd_rate is not None else None,
        "effective_zig_rate": _num(get_zig_rate(period)),
        "created_at": period.created_at.isoformat() if period.created_at else None,
        "updated_at": period.updated_at.isoformat() if period.updated_at else None,
    }
    if include_months:
        months = sorted(
            period.months.all(),
            key=lambda m: FY_MONTH_ORDER.index(m.month)
            if m.month in FY_MONTH_ORDER
            else 99,
        )
        data["months"] = [month_to_dict(m) for m in months]
        data["summary"] = compute_summary(months, get_zig_rate(period))
        data["subscriber_totals"] = compute_subscriber_totals(months)
    return data


def compute_summary(months: List[AkelloRevenueMonth], rate: Decimal) -> Dict[str, Any]:
    hlf_usd = Decimal("0")
    hlf_zwl = Decimal("0")
    org_usd = Decimal("0")
    org_zwl = Decimal("0")

    for m in months:
        hlf_usd += _dec(m.rev_asl_hlf_usd) + _dec(m.rev_lib_hlf_usd)
        hlf_zwl += _dec(m.rev_asl_hlf_zwl) + _dec(m.rev_lib_hlf_zwl)
        org_usd += _dec(m.rev_asl_org_usd) + _dec(m.rev_lib_org_usd)
        org_zwl += _dec(m.rev_asl_org_zwl) + _dec(m.rev_lib_org_zwl)

    rate = rate if rate > 0 else Decimal("37")
    hlf_zig = hlf_zwl / rate
    org_zig = org_zwl / rate
    hlf_total = hlf_usd + hlf_zig
    org_total = org_usd + org_zig
    grand_usd = hlf_usd + org_usd
    grand_zwl = hlf_zwl + org_zwl
    grand_zig = hlf_zig + org_zig
    grand_total = hlf_total + org_total

    def pct(part: Decimal) -> float:
        if grand_total <= 0:
            return 0.0
        return float(part / grand_total)

    return {
        "hlf": {
            "usd": _num(hlf_usd),
            "zwl": _num(hlf_zwl),
            "zig_usd": _num(hlf_zig),
            "total": _num(hlf_total),
        },
        "organic": {
            "usd": _num(org_usd),
            "zwl": _num(org_zwl),
            "zig_usd": _num(org_zig),
            "total": _num(org_total),
        },
        "total": {
            "usd": _num(grand_usd),
            "zwl": _num(grand_zwl),
            "zig_usd": _num(grand_zig),
            "total": _num(grand_total),
        },
        "contribution_pct": {
            "hlf": pct(hlf_total),
            "organic": pct(org_total),
            "total": 1.0 if grand_total > 0 else 0.0,
        },
        "zig_usd_rate": _num(rate),
        "note": f"***** rate of {_num(rate):g} was applied ZIG - USD",
    }


def compute_subscriber_totals(months: List[AkelloRevenueMonth]) -> Dict[str, int]:
    totals = {f: 0 for f in SUBSCRIBER_FIELDS}
    for m in months:
        for f in SUBSCRIBER_FIELDS:
            totals[f] += _int(getattr(m, f, 0))
    return totals


def apply_month_payload(row: AkelloRevenueMonth, data: Dict[str, Any]) -> None:
    for f in REVENUE_FIELDS:
        if f in data and data[f] is not None:
            setattr(row, f, _dec(data[f]))
        elif getattr(row, f, None) is None:
            setattr(row, f, Decimal("0"))
    for f in SUBSCRIBER_FIELDS:
        if f in data and data[f] is not None:
            setattr(row, f, _int(data[f]))
        elif getattr(row, f, None) is None:
            setattr(row, f, 0)


def seed_fy2027_if_empty() -> Dict[str, Any]:
    """Create FY2027 period and Mar–Jul rows if the period does not exist."""
    existing = AkelloRevenuePeriod.query.filter_by(code="FY2027").first()
    if existing:
        return {"created": False, "period_id": existing.id, "months": existing.months.count()}

    period = AkelloRevenuePeriod(code="FY2027", name="Financial Year 2027", zig_usd_rate=None)
    db.session.add(period)
    db.session.flush()

    for seed in FY2027_SEED_MONTHS:
        row = AkelloRevenueMonth(period_id=period.id, month=seed["month"])
        apply_month_payload(row, seed)
        db.session.add(row)

    db.session.commit()
    return {"created": True, "period_id": period.id, "months": len(FY2027_SEED_MONTHS)}


def find_period_by_code(code: str) -> Optional[AkelloRevenuePeriod]:
    code = (code or "").strip()
    if not code:
        return None
    period = AkelloRevenuePeriod.query.filter_by(code=code.upper()).first()
    if period:
        return period
    return AkelloRevenuePeriod.query.filter_by(code=code).first()


_MONTH_NAME_LOOKUP = {v.lower().strip(): k for k, v in MONTH_NAMES.items()}


def parse_month_label(val: Any) -> Optional[int]:
    if val is None:
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        m = int(val)
        return m if 1 <= m <= 12 else None
    text = str(val).strip()
    if not text or text.lower() == "total":
        return None
    if text.isdigit():
        m = int(text)
        return m if 1 <= m <= 12 else None
    key = text.lower().rstrip(".")
    return _MONTH_NAME_LOOKUP.get(key)


def _cell_num(val: Any) -> float:
    if val is None or val == "":
        return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def _cell_int(val: Any) -> int:
    return int(round(_cell_num(val)))


# Excel layout columns (1-indexed openpyxl): A=1 month, B–E HLF, G–J Organic
_REV_COL_MAP = {
    2: "rev_asl_hlf_usd",
    3: "rev_asl_hlf_zwl",
    4: "rev_lib_hlf_usd",
    5: "rev_lib_hlf_zwl",
    7: "rev_asl_org_usd",
    8: "rev_asl_org_zwl",
    9: "rev_lib_org_usd",
    10: "rev_lib_org_zwl",
}
_SUB_COL_MAP = {
    2: "sub_asl_hlf_usd",
    3: "sub_asl_hlf_zwl",
    4: "sub_lib_hlf_usd",
    5: "sub_lib_hlf_zwl",
    7: "sub_asl_org_usd",
    8: "sub_asl_org_zwl",
    9: "sub_lib_org_usd",
    10: "sub_lib_org_zwl",
}


def _parse_metric_sheet(ws, col_map: Dict[int, str], as_int: bool = False) -> tuple[Dict[int, Dict[str, Any]], List[str]]:
    rows: Dict[int, Dict[str, Any]] = {}
    errors: List[str] = []
    sheet_name = ws.title
    for r in range(4, (ws.max_row or 3) + 1):
        label = ws.cell(r, 1).value
        month = parse_month_label(label)
        if month is None:
            if label is not None and str(label).strip() and str(label).strip().lower() != "total":
                errors.append(f"{sheet_name} row {r}: unrecognized month '{label}'")
            continue
        payload: Dict[str, Any] = {"month": month}
        for col, field in col_map.items():
            raw = ws.cell(r, col).value
            payload[field] = _cell_int(raw) if as_int else _cell_num(raw)
        rows[month] = payload
    return rows, errors


def parse_akello_revenue_workbook(file_storage_or_path) -> Dict[str, Any]:
    """Parse Revenue + Subscribers sheets into month payloads."""
    from openpyxl import load_workbook

    wb = load_workbook(file_storage_or_path, data_only=True)
    names = {n.lower(): n for n in wb.sheetnames}
    errors: List[str] = []
    if "revenue" not in names:
        errors.append("Missing 'Revenue' sheet")
    if "subscribers" not in names:
        errors.append("Missing 'Subscribers' sheet")
    if errors:
        return {"months": [], "errors": errors}

    rev_rows, rev_errs = _parse_metric_sheet(wb[names["revenue"]], _REV_COL_MAP, as_int=False)
    sub_rows, sub_errs = _parse_metric_sheet(wb[names["subscribers"]], _SUB_COL_MAP, as_int=True)
    errors.extend(rev_errs)
    errors.extend(sub_errs)

    months: List[Dict[str, Any]] = []
    all_month_nums = sorted(
        set(rev_rows) | set(sub_rows),
        key=lambda m: FY_MONTH_ORDER.index(m) if m in FY_MONTH_ORDER else 99,
    )
    for m in all_month_nums:
        payload: Dict[str, Any] = {"month": m}
        if m in rev_rows:
            payload.update(rev_rows[m])
        else:
            errors.append(f"Month {MONTH_NAMES.get(m, m)} missing from Revenue sheet")
            for f in REVENUE_FIELDS:
                payload[f] = 0
        if m in sub_rows:
            payload.update(sub_rows[m])
        else:
            errors.append(f"Month {MONTH_NAMES.get(m, m)} missing from Subscribers sheet")
            for f in SUBSCRIBER_FIELDS:
                payload[f] = 0
        months.append(payload)

    return {"months": months, "errors": errors}


def apply_imported_months(
    period: AkelloRevenuePeriod,
    months: List[Dict[str, Any]],
    *,
    mode: str = "upsert",
    user_id: Optional[int] = None,
) -> Dict[str, Any]:
    """Upsert (or replace) month rows for a period from parsed payloads."""
    mode = (mode or "upsert").strip().lower()
    applied = 0
    imported_months = set()

    for payload in months:
        month = int(payload["month"])
        imported_months.add(month)
        row = AkelloRevenueMonth.query.filter_by(period_id=period.id, month=month).first()
        if not row:
            row = AkelloRevenueMonth(period_id=period.id, month=month)
            db.session.add(row)
        apply_month_payload(row, payload)
        row.updated_by = user_id
        applied += 1

    deleted = 0
    if mode == "replace":
        for row in list(period.months.all()):
            if row.month not in imported_months:
                db.session.delete(row)
                deleted += 1

    db.session.commit()
    return {"applied": applied, "deleted": deleted}


def _ordered_period_months(period: Optional[AkelloRevenuePeriod]) -> List[AkelloRevenueMonth]:
    if period is None:
        return []
    return sorted(
        period.months.all(),
        key=lambda m: FY_MONTH_ORDER.index(m.month) if m.month in FY_MONTH_ORDER else 99,
    )


def _write_metric_sheet_headers(ws) -> None:
    ws["B1"] = "Akello Smart Learning"
    ws["D1"] = "Akello Library"
    ws["G1"] = "Akello Smart Learning"
    ws["I1"] = "Akello Library"
    ws["B2"] = "HLF"
    ws["D2"] = "HLF"
    ws["G2"] = "Organic"
    ws["I2"] = "Organic"
    ws["A3"] = "Month"
    for col, label in [
        (2, "USD"),
        (3, "ZWL"),
        (4, "USD"),
        (5, "ZWL"),
        (7, "USD"),
        (8, "ZWL"),
        (9, "USD"),
        (10, "ZWL"),
    ]:
        ws.cell(3, col, label)


def _write_metric_sheet(
    wb,
    title: str,
    field_prefix: str,
    months_src: List[AkelloRevenueMonth],
    *,
    blank_starter: bool = False,
    include_totals: bool = False,
) -> None:
    ws = wb.create_sheet(title)
    _write_metric_sheet_headers(ws)
    mapping = _REV_COL_MAP if field_prefix == "rev" else _SUB_COL_MAP

    if not months_src and blank_starter:
        for i, month in enumerate(FY_MONTH_ORDER[:5], start=4):
            ws.cell(i, 1, MONTH_NAMES[month])
            for col in (2, 3, 4, 5, 7, 8, 9, 10):
                ws.cell(i, col, 0)
        return

    for i, row in enumerate(months_src, start=4):
        ws.cell(i, 1, MONTH_NAMES.get(row.month, row.month))
        for col, field in mapping.items():
            ws.cell(i, col, getattr(row, field, 0) or 0)

    if include_totals and months_src:
        total_row = 4 + len(months_src)
        ws.cell(total_row, 1, "Total")
        for col, field in mapping.items():
            total = 0
            for row in months_src:
                val = getattr(row, field, 0) or 0
                total += float(val) if field_prefix == "rev" else int(val)
            ws.cell(total_row, col, total if field_prefix == "rev" else int(total))


def _write_summary_sheet(wb, period: Optional[AkelloRevenuePeriod], months_src: List[AkelloRevenueMonth]) -> None:
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum["A1"] = "Period"
    ws_sum["B1"] = period.code if period else ""
    ws_sum["C1"] = period.name if period else ""
    ws_sum["A2"] = ""
    ws_sum["B3"] = "US$"
    ws_sum["C3"] = "ZWL"
    ws_sum["D3"] = "ZIG -USD$"
    ws_sum["E3"] = "Total"
    ws_sum["A4"] = "HLF Total"
    ws_sum["A5"] = "Organic"
    ws_sum["A6"] = "Total"
    ws_sum["A8"] = "% of Revenue contributions"
    ws_sum["A9"] = "HLF"
    ws_sum["A10"] = "Organic"
    ws_sum["A11"] = "Total"

    if period is not None and months_src:
        summary = compute_summary(months_src, get_zig_rate(period))
        for row_idx, key in [(4, "hlf"), (5, "organic"), (6, "total")]:
            block = summary.get(key) or {}
            ws_sum.cell(row_idx, 2, block.get("usd", 0))
            ws_sum.cell(row_idx, 3, block.get("zwl", 0))
            ws_sum.cell(row_idx, 4, block.get("zig_usd", 0))
            ws_sum.cell(row_idx, 5, block.get("total", 0))
        pct = summary.get("contribution_pct") or {}
        ws_sum["B9"] = float(pct.get("hlf") or 0)
        ws_sum["B10"] = float(pct.get("organic") or 0)
        ws_sum["B11"] = float(pct.get("total") or 0)
        ws_sum["A13"] = summary.get("note") or ""
    else:
        ws_sum["A13"] = "***** rate of 37 was applied ZIG -USD"


def build_period_template_bytes(period: Optional[AkelloRevenuePeriod] = None) -> bytes:
    """Build an .xlsx template matching the Revenue/Subscribers Excel layout."""
    from io import BytesIO

    from openpyxl import Workbook

    wb = Workbook()
    months_src = _ordered_period_months(period)
    _write_summary_sheet(wb, period, months_src)
    # Template keeps empty starter months when period has no data
    _write_metric_sheet(wb, "Revenue", "rev", months_src, blank_starter=not months_src, include_totals=False)
    _write_metric_sheet(wb, "Subscribers", "sub", months_src, blank_starter=not months_src, include_totals=False)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_period_report_bytes(period: AkelloRevenuePeriod) -> bytes:
    """Export selected FY period as it appears in the UI (Summary + Revenue + Subscribers)."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    months_src = _ordered_period_months(period)
    _write_summary_sheet(wb, period, months_src)
    _write_metric_sheet(wb, "Revenue", "rev", months_src, include_totals=True)
    _write_metric_sheet(wb, "Subscribers", "sub", months_src, include_totals=True)

    bold = Font(bold=True)
    for sheet_name in ("Summary", "Revenue", "Subscribers"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        if sheet_name == "Summary":
            for cell in ("A4", "A5", "A6", "A8", "A11", "B3", "C3", "D3", "E3"):
                ws[cell].font = bold
        else:
            for col in range(1, 11):
                ws.cell(3, col).font = bold
            # Total row
            if months_src:
                total_row = 4 + len(months_src)
                for col in range(1, 11):
                    if ws.cell(total_row, col).value is not None:
                        ws.cell(total_row, col).font = bold

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_fy_digest_html(period: AkelloRevenuePeriod) -> str:
    data = period_to_dict(period, include_months=True)
    summary = data.get("summary") or {}
    months = data.get("months") or []

    def money(n: Any) -> str:
        try:
            return f"{float(n):,.2f}"
        except (TypeError, ValueError):
            return "0.00"

    rows_html = "".join(
        f"<tr><td>{m.get('month_name')}</td>"
        f"<td style='text-align:right'>{money(m.get('rev_asl_hlf_usd'))}</td>"
        f"<td style='text-align:right'>{money(m.get('rev_lib_hlf_usd'))}</td>"
        f"<td style='text-align:right'>{money(m.get('rev_asl_org_usd'))}</td>"
        f"<td style='text-align:right'>{money(m.get('rev_lib_org_usd'))}</td></tr>"
        for m in months
    )
    hlf = summary.get("hlf") or {}
    org = summary.get("organic") or {}
    total = summary.get("total") or {}
    pct = summary.get("contribution_pct") or {}
    return f"""
    <h2>Akello Revenue Digest — {period.name} ({period.code})</h2>
    <p>{summary.get('note') or ''}</p>
    <h3>Summary</h3>
    <table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse;">
      <tr><th></th><th>US$</th><th>ZWL</th><th>ZIG-USD</th><th>Total</th></tr>
      <tr><td>HLF</td><td>{money(hlf.get('usd'))}</td><td>{money(hlf.get('zwl'))}</td>
          <td>{money(hlf.get('zig_usd'))}</td><td>{money(hlf.get('total'))}</td></tr>
      <tr><td>Organic</td><td>{money(org.get('usd'))}</td><td>{money(org.get('zwl'))}</td>
          <td>{money(org.get('zig_usd'))}</td><td>{money(org.get('total'))}</td></tr>
      <tr><td><b>Total</b></td><td><b>{money(total.get('usd'))}</b></td><td><b>{money(total.get('zwl'))}</b></td>
          <td><b>{money(total.get('zig_usd'))}</b></td><td><b>{money(total.get('total'))}</b></td></tr>
    </table>
    <p>Contribution: HLF {float(pct.get('hlf') or 0)*100:.1f}% · Organic {float(pct.get('organic') or 0)*100:.1f}%</p>
    <h3>Revenue by month (USD)</h3>
    <table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse;">
      <tr><th>Month</th><th>ASL HLF</th><th>Lib HLF</th><th>ASL Org</th><th>Lib Org</th></tr>
      {rows_html or '<tr><td colspan="5">No months</td></tr>'}
    </table>
    """


def run_akello_revenue_digest(*, triggered_by: str = "scheduler", period_code: Optional[str] = None) -> Dict[str, Any]:
    """Email FY revenue digest to configured recipients."""
    from app.email_utils import send_html_email
    from app.revenue_email_utils import resolve_revenue_report_recipients

    enabled = (AppSetting.get_value("akello_revenue_digest_enabled", "false") or "false").strip().lower() in (
        "true",
        "1",
        "yes",
        "y",
        "t",
    )
    if not enabled and triggered_by == "scheduler":
        return {"status": "skipped", "reason": "digest_disabled"}

    seed_fy2027_if_empty()
    code = (period_code or AppSetting.get_value("akello_revenue_digest_period", "FY2027") or "FY2027").strip()
    period = find_period_by_code(code)
    if not period:
        return {"status": "failed", "reason": f"period_not_found:{code}"}

    recipients, reason = resolve_revenue_report_recipients()
    if not recipients:
        return {"status": "skipped", "reason": reason or "no_recipients"}

    html = build_fy_digest_html(period)
    subject = f"Akello Revenue Digest — {period.code}"
    sent = 0
    errors: List[str] = []
    for email in recipients:
        try:
            send_html_email(to_email=email, subject=subject, html_body=html)
            sent += 1
        except Exception as exc:  # noqa: BLE001
            errors.append(f"{email}: {exc}")

    status = "success" if sent and not errors else ("partial" if sent else "failed")
    AppSetting.set_value("akello_revenue_digest_last_status", status)
    AppSetting.set_value("akello_revenue_digest_last_reason", "; ".join(errors) if errors else reason or "")
    AppSetting.set_value("akello_revenue_digest_last_at", __import__("datetime").datetime.utcnow().isoformat())
    AppSetting.set_value("akello_revenue_digest_last_triggered_by", triggered_by)
    return {
        "status": status,
        "sent": sent,
        "recipients": recipients,
        "errors": errors,
        "period": period.code,
        "triggered_by": triggered_by,
    }
